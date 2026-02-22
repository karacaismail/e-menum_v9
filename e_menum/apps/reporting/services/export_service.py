"""
Export Service for generating report files in various formats.

Supports:
    - PDF: Report-quality PDF with tables and charts (ReportLab)
    - Excel: Multi-sheet workbooks with formatting (openpyxl)
    - CSV: Simple tabular data export (stdlib csv)

Usage:
    service = ExportService()

    # Export to Excel
    file_bytes = service.export_to_excel(report_data)

    # Export to CSV
    file_bytes = service.export_to_csv(report_data)

    # Export to PDF
    file_bytes = service.export_to_pdf(report_data)
"""

import csv
import io
import logging
from datetime import datetime
from decimal import Decimal
from typing import Any, Dict, List, Optional

from django.utils import timezone

logger = logging.getLogger(__name__)


class ExportService:
    """
    Service for exporting report data to various file formats.
    """

    def export(self, report_data: dict, format: str) -> bytes:
        """
        Export report data to the specified format.

        Args:
            report_data: The report result data
            format: Export format ('PDF', 'EXCEL', 'CSV', 'JSON')

        Returns:
            bytes: The exported file content

        Raises:
            ValueError: If format is not supported
        """
        format_upper = format.upper()

        if format_upper == 'EXCEL':
            return self.export_to_excel(report_data)
        elif format_upper == 'CSV':
            return self.export_to_csv(report_data)
        elif format_upper == 'PDF':
            return self.export_to_pdf(report_data)
        elif format_upper == 'JSON':
            return self.export_to_json(report_data)
        else:
            raise ValueError(f'Unsupported export format: {format}')

    def export_to_json(self, report_data: dict) -> bytes:
        """
        Export report data as formatted JSON.

        Args:
            report_data: The report result data

        Returns:
            bytes: JSON content
        """
        import json

        class DecimalEncoder(json.JSONEncoder):
            def default(self, obj):
                if isinstance(obj, Decimal):
                    return float(obj)
                if isinstance(obj, datetime):
                    return obj.isoformat()
                return super().default(obj)

        content = json.dumps(
            report_data,
            cls=DecimalEncoder,
            indent=2,
            ensure_ascii=False,
        )
        return content.encode('utf-8')

    def export_to_csv(self, report_data: dict) -> bytes:
        """
        Export report data as CSV.

        Expects report_data to contain a 'rows' key with list of dicts,
        or a 'data' key with list of dicts.

        Args:
            report_data: The report result data

        Returns:
            bytes: CSV content
        """
        output = io.StringIO()

        # Extract tabular data
        rows = self._extract_rows(report_data)

        if not rows:
            output.write('No data available\n')
            return output.getvalue().encode('utf-8')

        # Write CSV
        writer = csv.DictWriter(output, fieldnames=rows[0].keys())
        writer.writeheader()
        for row in rows:
            # Convert non-string values
            clean_row = {}
            for k, v in row.items():
                if isinstance(v, Decimal):
                    clean_row[k] = float(v)
                elif isinstance(v, (dict, list)):
                    clean_row[k] = str(v)
                else:
                    clean_row[k] = v
            writer.writerow(clean_row)

        return output.getvalue().encode('utf-8')

    def export_to_excel(self, report_data: dict) -> bytes:
        """
        Export report data as Excel workbook.

        Creates a formatted workbook with:
            - Summary sheet with KPI metrics
            - Data sheet with tabular data
            - Charts sheet (if applicable)

        Args:
            report_data: The report result data

        Returns:
            bytes: Excel file content (.xlsx)
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            logger.error('openpyxl not installed. Install with: pip install openpyxl')
            raise ImportError('openpyxl is required for Excel export')

        wb = Workbook()

        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = 'Summary'

        # Title styling
        title_font = Font(name='Calibri', size=14, bold=True)
        header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')

        # Write summary
        ws_summary['A1'] = report_data.get('title', 'Report')
        ws_summary['A1'].font = title_font
        ws_summary['A2'] = f'Generated: {timezone.now().strftime("%Y-%m-%d %H:%M")}'

        # Write summary metrics
        summary = report_data.get('summary', {})
        row = 4
        for key, value in summary.items():
            ws_summary.cell(row=row, column=1, value=str(key))
            ws_summary.cell(row=row, column=1).font = Font(bold=True)
            ws_summary.cell(row=row, column=2, value=self._format_value(value))
            row += 1

        # Data sheet
        rows = self._extract_rows(report_data)
        if rows:
            ws_data = wb.create_sheet('Data')

            # Headers
            headers = list(rows[0].keys())
            for col_idx, header in enumerate(headers, 1):
                cell = ws_data.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')

            # Data rows
            for row_idx, row_data in enumerate(rows, 2):
                for col_idx, header in enumerate(headers, 1):
                    value = row_data.get(header, '')
                    ws_data.cell(
                        row=row_idx,
                        column=col_idx,
                        value=self._format_value(value),
                    )

            # Auto-width columns
            for col_idx, header in enumerate(headers, 1):
                max_width = max(
                    len(str(header)),
                    max(
                        (len(str(row_data.get(header, ''))) for row_data in rows),
                        default=0,
                    ),
                )
                ws_data.column_dimensions[
                    get_column_letter(col_idx)
                ].width = min(max_width + 2, 50)

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    def export_to_pdf(self, report_data: dict) -> bytes:
        """
        Export report data as PDF.

        Creates a professional PDF report with:
            - Header with report title and date
            - Summary metrics section
            - Data table
            - Footer with page numbers

        Args:
            report_data: The report result data

        Returns:
            bytes: PDF file content
        """
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
            from reportlab.lib.units import cm, mm
            from reportlab.platypus import (
                Paragraph,
                SimpleDocTemplate,
                Spacer,
                Table,
                TableStyle,
            )
        except ImportError:
            logger.error('reportlab not installed. Install with: pip install reportlab')
            raise ImportError('reportlab is required for PDF export')

        output = io.BytesIO()
        doc = SimpleDocTemplate(
            output,
            pagesize=A4,
            rightMargin=2 * cm,
            leftMargin=2 * cm,
            topMargin=2 * cm,
            bottomMargin=2 * cm,
        )

        styles = getSampleStyleSheet()
        elements = []

        # Title
        title = report_data.get('title', 'Report')
        title_style = ParagraphStyle(
            'ReportTitle',
            parent=styles['Title'],
            fontSize=18,
            spaceAfter=12,
        )
        elements.append(Paragraph(title, title_style))
        elements.append(Paragraph(
            f'Generated: {timezone.now().strftime("%Y-%m-%d %H:%M")}',
            styles['Normal'],
        ))
        elements.append(Spacer(1, 12))

        # Summary section
        summary = report_data.get('summary', {})
        if summary:
            elements.append(Paragraph('Summary', styles['Heading2']))
            summary_data = [[str(k), str(self._format_value(v))] for k, v in summary.items()]
            if summary_data:
                t = Table(summary_data, colWidths=[8 * cm, 8 * cm])
                t.setStyle(TableStyle([
                    ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                    ('TOPPADDING', (0, 0), (-1, -1), 6),
                ]))
                elements.append(t)
                elements.append(Spacer(1, 12))

        # Data table
        rows = self._extract_rows(report_data)
        if rows:
            elements.append(Paragraph('Data', styles['Heading2']))
            headers = list(rows[0].keys())

            table_data = [headers]
            for row in rows[:100]:  # Limit to 100 rows for PDF
                table_data.append([
                    str(self._format_value(row.get(h, '')))
                    for h in headers
                ])

            # Calculate column widths
            num_cols = len(headers)
            col_width = min(16 * cm / num_cols, 5 * cm)

            t = Table(table_data, colWidths=[col_width] * num_cols)
            t.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563EB')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F3F4F6')]),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
            ]))
            elements.append(t)

            if len(rows) > 100:
                elements.append(Spacer(1, 6))
                elements.append(Paragraph(
                    f'Showing 100 of {len(rows)} rows. Export to Excel for full data.',
                    styles['Italic'],
                ))

        doc.build(elements)
        return output.getvalue()

    # =========================================================================
    # HELPERS
    # =========================================================================

    def _extract_rows(self, report_data: dict) -> List[dict]:
        """
        Extract tabular rows from report data.

        Tries keys: 'rows', 'data', 'items', 'results'

        Args:
            report_data: The report result data

        Returns:
            list[dict]: List of row dicts
        """
        for key in ('rows', 'data', 'items', 'results', 'products', 'orders'):
            if key in report_data and isinstance(report_data[key], list):
                rows = report_data[key]
                if rows and isinstance(rows[0], dict):
                    return rows
        return []

    def _format_value(self, value: Any) -> Any:
        """
        Format a value for display in exports.

        Args:
            value: The raw value

        Returns:
            Formatted value
        """
        if isinstance(value, Decimal):
            return float(value)
        if isinstance(value, datetime):
            return value.strftime('%Y-%m-%d %H:%M')
        if isinstance(value, (dict, list)):
            return str(value)
        return value
